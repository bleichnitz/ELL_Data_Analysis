import _load_files as l_files
import _clean_data as c_data
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# import data from our working ELL list
WB_WORKING_ELL_LIST = l_files.load_file("ELL List (updated 2022-10-31 ).xlsx")
sheet_name = WB_WORKING_ELL_LIST.sheetnames
sheet_ELL_LIST = WB_WORKING_ELL_LIST[sheet_name[0]]
ELL_LIST = c_data.clean_ell_list(data=sheet_ELL_LIST)
clean_list_sheet = WB_WORKING_ELL_LIST["Clean List"]

# import data from most recent SIS ELL list
WB_SIS_ELL_LIST = l_files.load_file("2022-11-18 ELL List.xlsx")
sheet_name = WB_SIS_ELL_LIST.sheetnames
sheet_SIS_ELL_LIST = WB_SIS_ELL_LIST[sheet_name[0]]
SIS_ELL_LIST = c_data.clean_SIS_ell_list(data=sheet_SIS_ELL_LIST)

# c_data.add_remove_students(working_list=ELL_LIST, sis_list=SIS_ELL_LIST)

# import data from master grade summary
WB_MASTER_GRADE_SUMMARY = l_files.load_file("2022-11-18 Master Grade Summary2.xlsx")
sheet_name = WB_MASTER_GRADE_SUMMARY.sheetnames
sheet_MASTER_GRADE_SUMMARY = WB_MASTER_GRADE_SUMMARY[sheet_name[0]]
"""MASTER_GRADE_SUMMARY = c_data.clean_grade_summary(grade_data=sheet_MASTER_GRADE_SUMMARY, ell_list=ELL_LIST)

# print grade data to spreadsheet
print("")
print("")
for i in range(0, len(ELL_LIST)):
    student_data = []
    current_student = ELL_LIST[i][0]
    r = 5 + i
    for s in range(0, len(MASTER_GRADE_SUMMARY)):
        if current_student == MASTER_GRADE_SUMMARY[s][0][0]:
            student_data = MASTER_GRADE_SUMMARY[s]
            cell = clean_list_sheet.cell(row=r, column=22)
            avg = student_data[0][1]
            cell.value = avg
            cell.alignment = Alignment(horizontal='center', vertical='center')
            for c in range(0,len(student_data)):
                grade = student_data[c][3]
                cell = clean_list_sheet.cell(row=r, column=23+c)
                cell.value = grade
                cell.alignment = Alignment(horizontal='center', vertical='center')
WB_WORKING_ELL_LIST.save("ELL List (updated 2022-10-31 ).xlsx")
"""



# PRINT LANGUAGE
for s in range(0, len(ELL_LIST)):
    student_num = ELL_LIST[s][0]
    student_data = c_data.filter_for_student(student=student_num, list=SIS_ELL_LIST)
    print(student_data)
    sex = student_data[0][3]
    cell = clean_list_sheet.cell(row=5+s, column=4)
    cell.value = sex
    cell.alignment = Alignment(horizontal='center', vertical='center')

# PRINT GRADE
for s in range(0, len(ELL_LIST)):
    student_num = ELL_LIST[s][0]
    student_data = c_data.filter_for_student(student=student_num, list=SIS_ELL_LIST)
    print(student_data)
    sex = student_data[0][4]
    cell = clean_list_sheet.cell(row=5+s, column=5)
    cell.value = sex
    cell.alignment = Alignment(horizontal='center', vertical='center')

#PRINT L1
for s in range(0, len(ELL_LIST)):
    student_num = ELL_LIST[s][0]
    student_data = c_data.filter_for_student(student=student_num, list=SIS_ELL_LIST)
    print(student_data)
    sex = student_data[0][5]
    cell = clean_list_sheet.cell(row=5+s, column=6)
    cell.value = sex
    cell.alignment = Alignment(horizontal='center', vertical='center')

#PRINT L2
for s in range(0, len(ELL_LIST)):
    student_num = ELL_LIST[s][0]
    student_data = c_data.filter_for_student(student=student_num, list=SIS_ELL_LIST)
    print(student_data)
    sex = student_data[0][6]
    cell = clean_list_sheet.cell(row=5+s, column=7)
    cell.value = sex
    cell.alignment = Alignment(horizontal='center', vertical='center')

WB_WORKING_ELL_LIST.save("ELL List (updated 2022-10-31 ).xlsx")
